import serial
import time
from openpyxl import Workbook, load_workbook
"""from facerec import verify_access
from scanner import read_barcode"""
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pyrebase

# Firebase configuration
firebaseConfig = {
    "apiKey": "AIzaSyCgcQ5m5xMxA0mnPEAeFU9dXQXSmt6cZtU",
    "authDomain": "cloud-integration-c6dde.firebaseapp.com",
    "databaseURL": "https://cloud-integration-c6dde-default-rtdb.asia-southeast1.firebasedatabase.app",
    "projectId": "cloud-integration-c6dde",
    "storageBucket": "cloud-integration-c6dde.firebasestorage.app",
    "messagingSenderId": "89955976109",
    "appId": "1:89955976109:web:e938bd8f6ff612ffe55724",
    "measurementId": "G-E344YC5WYG"
}

firebase = pyrebase.initialize_app(firebaseConfig)
db = firebase.database()

ser = serial.Serial('COM5', 9600, timeout=1)  
time.sleep(2)

item_database = "items.xlsx"
database_file = "DATABASE.xlsx"
try:
    db_workbook = load_workbook(database_file)
    db_sheet = db_workbook.active
    idb_workbook = load_workbook(item_database)
    idb_sheet = idb_workbook.active
except FileNotFoundError:
    print("Database file not found. Please check the file path.")
    exit()

scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("studious-pulsar-451107-d2-3508324c506e.json", scope)  
client = gspread.authorize(creds)

sheet = client.open("serial_database").sheet1

print("Waiting for data...")

def find_in_database(card_sn):
    for row in db_sheet.iter_rows(min_row=2, values_only=True):  
        if row[0] == card_sn:
            name, reg_no = row[1], row[2]
            print(f"RFID Detected: {reg_no}, {name}")
            verified_user = verify_access()
            if verified_user != name:
                print("Face Mismatch ❌ Access Denied")
                return None, None, None
            print(f"Face Verified ✅ Access Granted for {name}")

            items = []
            while True:
                print("Scan an item barcode and qrcode")
                item = read_barcode().strip()
                op=int(input("Enter your choice: "))
            
                for i in idb_sheet.iter_rows(min_row=2, values_only=True):
                    if i[0] == item:
                        items.append([i[1],op,item])

                
                n = int(input("1. Continue scanning\n2. Finish\nEnter : "))
                if n == 2:
                    break

            return reg_no, name, items

    return None, None, None

def find_row(value1, value2):
    row=1
    v1=sheet.cell(1,4).value
    v2=sheet.cell(1,7).value
    while sheet.cell(row,4).value and sheet.cell(row,7).value:
        print(sheet.cell(row,4).value,sheet.cell(row,7).value)
        if sheet.cell(row,4).value==value1 and sheet.cell(row,7).value==value2:
            return row
        row+=1

    print(f"❌ No matching row found for values ({value1}, {value2})")  
    return None

try:
    while True:
        if ser.in_waiting > 0:  
            data = ser.readline().decode('utf-8').strip().split()
            data=data[-1]
            print(f"Received: {data}")
            
            reg_no, name, items = find_in_database(data)
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            
            if reg_no and name:
                for item in items:
                    firebase_data = {
                        "timestamp": timestamp,
                        "name": name,
                        "rfid": reg_no,
                        "qr_code": item[2],
                        "action": "Check-in" if item[1] == 1 else "Check-out"
                    }
                    db.child("lab_entries").push(firebase_data)
                    print(f"Data uploaded to Firebase: {firebase_data}")
                    
                    if item[1] == 1:
                        row = [timestamp, None, name, reg_no, item[0], data, item[2]]
                        sheet.append_row(row)
                    if item[1] == 2:
                        print("2")
                        col = 2
                        row = find_row(reg_no, item[2])

                        if row:
                            sheet.update_cell(row, col, timestamp)
                            print(f"✅ Updated timestamp in row {row}, column {col}")
                        else:
                            print(f"❌ Could not update Google Sheet - No row found for ({reg_no}, {item[2]})")

                print("Data written to Google Sheet and Firebase")
            else:
                print(f"Error: Card SN '{data}' not found in database.")

except KeyboardInterrupt:
    print("Program stopped manually")

finally:
    ser.close()
    print("File saved.")
